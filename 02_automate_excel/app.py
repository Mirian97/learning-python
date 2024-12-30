import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# Ways to access a cell within workbook
# cell = sheet["a1"]
# cell = sheet.cell(1,1)

def proccess_workbook (filename: str):
    # Load the specified Excel workbook.
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]  # Select the worksheet named "Sheet1".
    max_rows = sheet.max_row  # Get the total number of used rows in the worksheet.

    # Iterate through the rows, starting from row 2 (skipping the headers).
    for row in range(2, max_rows + 1):
        cell_price = sheet.cell(row, 3)  # Get the value in column 3 (original price).
        corrected_price = cell_price.value * 0.9  # Apply a 10% discount to the price.
        corrected_price_cell = sheet.cell(row, 4)  # Select the cell in column 4 for the corrected price.
        corrected_price_cell.value = corrected_price  # Insert the corrected price into the cell.

    # Create a reference to the values in column 4 (corrected prices) for the chart.
    reference_values = Reference(
        sheet,
        min_row=2,  # Start from row 2 (skip the header).
        max_row=max_rows,  # Go up to the last filled row.
        max_col=4,  # Select column 4 (corrected prices).
        min_col=4  # Select the same column.
    )
    chart = BarChart()  # Create a bar chart.
    chart.add_data(reference_values)  # Add the reference data to the chart.
    sheet.add_chart(chart, "e2")  # Place the chart at position "E2" in the worksheet.

    # Save the changes to the Excel file.
    wb.save(filename)
    
proccess_workbook("transactions.xlsx")